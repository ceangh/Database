import openpyxl
from colorama import Fore, Style

# Pfad zur Excel-Datei
excel_path = r'O:\REDAKTION\Herstellerdokus\Doku Artikel DB - Kopie.xlsm'

# Funktion zum Suchen nach Ordernummer
def search_by_ordernumber(ordernumber):
    # Öffne die Excel-Datei
    wb = openpyxl.load_workbook(excel_path)
    # Wähle das Arbeitsblatt aus (angenommen, es heißt 'DB', passe es an, wenn es anders ist)
    sheet = wb['DB']

    # Durchsuche die Spalte mit den Ordernummern
    for row in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row, column=1).value) == str(ordernumber):  # Vergleich als Zeichenketten
            # Wenn die Ordernummer gefunden wurde, gib die Informationen aus
            print(f'Ordernummer: {ordernumber}')
            print(f'Type: {sheet.cell(row=row, column=2).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'Manufacturer Folder: {sheet.cell(row=row, column=3).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'CDS_DE: {sheet.cell(row=row, column=4).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'CDS_EN: {sheet.cell(row=row, column=5).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'MANUAL_DE: {sheet.cell(row=row, column=6).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'MANUAL_EN: {sheet.cell(row=row, column=7).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'CERT_DoC: {sheet.cell(row=row, column=8).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'CERT_ATEX: {sheet.cell(row=row, column=9).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'CERT_MAR: {sheet.cell(row=row, column=10).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'CERT_CSA: {sheet.cell(row=row, column=11).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'CERT_UL: {sheet.cell(row=row, column=12).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'CSA/C-UL Category No: {sheet.cell(row=row, column=13).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'CSA/c-UL File No: {sheet.cell(row=row, column=14).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'UL_Category_No: {sheet.cell(row=row, column=15).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            print(f'UL_File_No: {sheet.cell(row=row, column=16).value or Fore.RED + "Empty" + Style.RESET_ALL}')
            # Füge hier weitere Spalten hinzu, die du benötigst

            # Schließe die Excel-Datei
            wb.close()
            return

    # Falls die Ordernummer nicht gefunden wurde
    print(Fore.RED + f'Ordernummer {ordernumber} nicht gefunden.' + Style.RESET_ALL)

    # Menüoptionen
    while True:
        print("\nMenü:")
        print("1. Erstellen")
        print("2. Ergänzen")
        print("3. Fertig")

        # Benutzerwahl abfragen
        choice = input("Wähle eine Option (1, 2, 3): ")

        if choice == "1":
            # Logik für Option 1 (Erstellen) hier einfügen
            print("Du hast Option 1 gewählt.")
        elif choice == "2":
            # Logik für Option 2 (Ergänzen) hier einfügen
            print("Du hast Option 2 gewählt.")
        elif choice == "3":
            print("Programm beendet.")
            break
        else:
            print("Ungültige Eingabe. Bitte wähle 1, 2 oder 3.")
# Frag nach der Ordernummer
ordernumber = input('Gib die Ordernummer ein: ')
search_by_ordernumber(ordernumber)
